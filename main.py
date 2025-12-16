from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func
from contextlib import asynccontextmanager
import uuid
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 导入数据库配置
from database import get_db, init_db, QuizRecord, UserScore, QuizStat

# 使用 lifespan 管理启动和关闭事件
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 启动时初始化数据库
    init_db()
    print("数据库已初始化")
    yield
    # 关闭时的清理工作（如果需要）
    print("应用关闭")

app = FastAPI(
    title="听音审美知识专栏 API",
    version="1.0.0",
    lifespan=lifespan
)

# CORS 配置 - 允许前端访问
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 生产环境应该指定具体域名
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== 数据模型 ====================

class Comment(BaseModel):
    """评论模型"""
    id: str = Field(default_factory=lambda: f"comment_{uuid.uuid4().hex[:12]}")
    userName: str
    content: str
    time: str
    likes: int = 0
    liked: bool = False

class CommentCreate(BaseModel):
    """创建评论请求模型"""
    userName: str
    content: str

class CommentLike(BaseModel):
    """点赞请求模型"""
    commentId: str
    liked: bool

class QuizOption(BaseModel):
    """问题选项模型"""
    label: str
    text: str
    isCorrect: bool

class Quiz(BaseModel):
    """互动问题模型"""
    id: str
    question: str
    options: List[QuizOption]

class QuizAnswer(BaseModel):
    """用户答案模型"""
    quizId: str
    selectedOption: str
    userName: Optional[str] = None
    userId: Optional[str] = None

class QuizResult(BaseModel):
    """答题结果模型"""
    isCorrect: bool
    correctAnswer: str
    message: str

# ==================== 内存数据存储 ====================
# 生产环境应该使用数据库

comments_db: List[Comment] = []

quizzes_db = {
    "quiz_1": Quiz(
        id="quiz_1",
        question="在进行日常听音训练或混音练习时，为什么建议将重放声压级控制在 75–85 dB 之间？",
        options=[
            QuizOption(label="A", text="因为这个声压级既能保持听觉敏感度，又能防止听力疲劳", isCorrect=True),
            QuizOption(label="B", text="因为声音越大越容易听清细节", isCorrect=False),
            QuizOption(label="C", text="因为小音量下听不出差别", isCorrect=False),
            QuizOption(label="D", text="因为这个范围最接近日常真实聆听环境，能帮助培养自然听感", isCorrect=False),
        ]
    )
}

# 用户答题记录：{userId: {quizId: [answer_records]}}
user_quiz_records = {}

# 用户积分统计：{userName: {correct: int, wrong: int, score: int}}
user_scores = {}

# 问题统计：{quizId: {correct: int, wrong: int}}
quiz_stats = {}

# ==================== 工具函数 ====================

def format_time() -> str:
    """格式化当前时间为 HH:MM"""
    now = datetime.now()
    return now.strftime("%H:%M")

# ==================== 评论相关 API ====================

@app.post("/api/comments", response_model=Comment)
async def create_comment(comment_data: CommentCreate):
    """创建新评论"""
    comment = Comment(
        userName=comment_data.userName,
        content=comment_data.content,
        time=format_time()
    )
    comments_db.insert(0, comment)  # 新评论插入到最前面
    return comment

@app.get("/api/comments", response_model=List[Comment])
async def get_comments():
    """获取所有评论"""
    return comments_db

@app.post("/api/comments/like")
async def toggle_like(like_data: CommentLike):
    """切换评论点赞状态"""
    comment = next((c for c in comments_db if c.id == like_data.commentId), None)

    if not comment:
        raise HTTPException(status_code=404, detail="评论不存在")

    if like_data.liked:
        comment.likes += 1
        comment.liked = True
    else:
        comment.likes = max(0, comment.likes - 1)
        comment.liked = False

    return {"success": True, "likes": comment.likes}

@app.delete("/api/comments/{comment_id}")
async def delete_comment(comment_id: str):
    """删除评论（可选功能）"""
    global comments_db
    comments_db = [c for c in comments_db if c.id != comment_id]
    return {"success": True}

# ==================== 互动问题相关 API ====================

@app.get("/api/quizzes/{quiz_id}", response_model=Quiz)
async def get_quiz(quiz_id: str):
    """获取问题详情"""
    quiz = quizzes_db.get(quiz_id)
    if not quiz:
        raise HTTPException(status_code=404, detail="问题不存在")
    return quiz

@app.get("/api/quizzes", response_model=List[Quiz])
async def get_all_quizzes():
    """获取所有问题"""
    return list(quizzes_db.values())

@app.post("/api/quizzes/answer", response_model=QuizResult)
async def submit_answer(answer: QuizAnswer, db: Session = Depends(get_db)):
    """提交答案并返回结果（使用数据库存储）"""
    quiz = quizzes_db.get(answer.quizId)

    if not quiz:
        raise HTTPException(status_code=404, detail="问题不存在")

    # 查找正确答案
    correct_option = next((opt for opt in quiz.options if opt.isCorrect), None)
    selected_option = next((opt for opt in quiz.options if opt.label == answer.selectedOption), None)

    if not correct_option or not selected_option:
        raise HTTPException(status_code=400, detail="无效的选项")

    is_correct = selected_option.isCorrect
    message = "回答正确！" if is_correct else "回答错误，重新试试吧"

    # 记录用户答题记录到数据库
    if answer.userName and answer.userId:
        # 1. 保存答题记录
        quiz_record = QuizRecord(
            quiz_id=answer.quizId,
            user_id=answer.userId,
            user_name=answer.userName,
            selected_option=answer.selectedOption,
            is_correct=is_correct
        )
        db.add(quiz_record)

        # 2. 更新或创建用户积分
        user_score = db.query(UserScore).filter(UserScore.user_name == answer.userName).first()

        if not user_score:
            # 创建新用户积分记录
            user_score = UserScore(
                user_name=answer.userName,
                correct_count=1 if is_correct else 0,
                wrong_count=0 if is_correct else 1,
                total_score=8 if is_correct else 0
            )
            db.add(user_score)
        else:
            # 更新现有用户积分
            if is_correct:
                user_score.correct_count += 1
                user_score.total_score += 8
            else:
                user_score.wrong_count += 1

        # 3. 更新问题统计
        quiz_stat = db.query(QuizStat).filter(QuizStat.quiz_id == answer.quizId).first()

        if not quiz_stat:
            # 创建新问题统计记录
            quiz_stat = QuizStat(
                quiz_id=answer.quizId,
                correct_count=1 if is_correct else 0,
                wrong_count=0 if is_correct else 1,
                total_count=1
            )
            db.add(quiz_stat)
        else:
            # 更新现有问题统计
            if is_correct:
                quiz_stat.correct_count += 1
            else:
                quiz_stat.wrong_count += 1
            quiz_stat.total_count += 1

        # 提交所有更改
        db.commit()

    return QuizResult(
        isCorrect=is_correct,
        correctAnswer=correct_option.label,
        message=message
    )

@app.get("/api/quizzes/{quiz_id}/stats")
async def get_quiz_stats(quiz_id: str):
    """获取问题答题统计"""
    stats = quiz_stats.get(quiz_id, {"correct": 0, "wrong": 0})
    total = stats["correct"] + stats["wrong"]
    accuracy = (stats["correct"] / total * 100) if total > 0 else 0

    return {
        **stats,
        "total": total,
        "accuracy": round(accuracy, 2)
    }

# ==================== 数据统计 API ====================

@app.get("/api/stats/users")
async def get_user_stats(db: Session = Depends(get_db)):
    """获取所有用户积分排行（从数据库读取）"""
    # 从数据库查询，按积分排序
    users = db.query(UserScore).order_by(UserScore.total_score.desc()).all()

    result = []
    for rank, user in enumerate(users, 1):
        total = user.correct_count + user.wrong_count
        accuracy = round(user.correct_count / total * 100, 2) if total > 0 else 0

        result.append({
            "rank": rank,
            "userName": user.user_name,
            "score": user.total_score,
            "correct": user.correct_count,
            "wrong": user.wrong_count,
            "total": total,
            "accuracy": accuracy
        })

    return result

@app.get("/api/stats/overview")
async def get_overview_stats(db: Session = Depends(get_db)):
    """获取总体数据统计（从数据库读取）"""
    # 统计用户数
    total_users = db.query(func.count(UserScore.id)).scalar()

    # 统计答题记录
    total_answers = db.query(func.count(QuizRecord.id)).scalar()
    total_correct = db.query(func.count(QuizRecord.id)).filter(QuizRecord.is_correct == True).scalar()

    # 评论数（仍使用内存）
    total_comments = len(comments_db)

    return {
        "totalUsers": total_users,
        "totalComments": total_comments,
        "totalAnswers": total_answers,
        "totalCorrect": total_correct,
        "overallAccuracy": round(total_correct / total_answers * 100, 2) if total_answers > 0 else 0
    }

@app.get("/api/stats/comments")
async def get_comment_stats():
    """获取评论统计"""
    # 统计每个用户的评论数
    user_comment_count = {}
    for comment in comments_db:
        userName = comment.userName
        user_comment_count[userName] = user_comment_count.get(userName, 0) + 1

    # 按评论数排序
    sorted_comments = sorted(
        user_comment_count.items(),
        key=lambda x: x[1],
        reverse=True
    )

    return [
        {"userName": userName, "commentCount": count}
        for userName, count in sorted_comments
    ]

@app.get("/api/stats/quiz-records")
async def get_all_quiz_records(db: Session = Depends(get_db)):
    """获取所有答题记录（从数据库读取）"""
    # 从数据库查询，按时间倒序
    records = db.query(QuizRecord).order_by(QuizRecord.answered_at.desc()).all()

    all_records = []
    for record in records:
        all_records.append({
            "userName": record.user_name,
            "quizId": record.quiz_id,
            "selectedOption": record.selected_option,
            "isCorrect": record.is_correct,
            "time": record.answered_at.strftime("%H:%M"),
            "timestamp": record.answered_at.isoformat()
        })

    return all_records

# ==================== Excel导出功能 ====================

def export_quiz_records_to_excel(db: Session):
    """导出答题记录到Excel文件（从数据库读取）"""
    # 确保答题情况目录存在（云端使用临时目录）
    export_dir = os.path.join(os.getcwd(), "答题情况")
    os.makedirs(export_dir, exist_ok=True)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "答题记录"

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    cell_font = Font(name='微软雅黑', size=11)
    cell_alignment = Alignment(horizontal='center', vertical='center')

    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # 写入表头
    headers = ['序号', '用户名', '问题ID', '选择答案', '答题结果', '答题时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 从数据库获取所有答题记录，按时间倒序
    records = db.query(QuizRecord).order_by(QuizRecord.answered_at.desc()).all()

    all_records = []
    for record in records:
        all_records.append({
            "userName": record.user_name,
            "quizId": record.quiz_id,
            "selectedOption": record.selected_option,
            "isCorrect": record.is_correct,
            "time": record.answered_at.strftime("%H:%M"),
            "timestamp": record.answered_at.isoformat()
        })

    # 写入数据
    for idx, record in enumerate(all_records, 2):
        # 序号
        cell = ws.cell(row=idx, column=1, value=idx-1)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border

        # 用户名
        cell = ws.cell(row=idx, column=2, value=record["userName"])
        cell.font = Font(name='微软雅黑', size=11, bold=True)
        cell.alignment = cell_alignment
        cell.border = border

        # 问题ID
        cell = ws.cell(row=idx, column=3, value=record["quizId"])
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border

        # 选择答案
        cell = ws.cell(row=idx, column=4, value=record["selectedOption"])
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border

        # 答题结果
        result_text = "✓ 正确" if record["isCorrect"] else "✗ 错误"
        cell = ws.cell(row=idx, column=5, value=result_text)
        cell.font = Font(name='微软雅黑', size=11, bold=True,
                        color='008000' if record["isCorrect"] else 'FF0000')
        cell.alignment = cell_alignment
        cell.border = border

        # 根据结果设置背景色
        if record["isCorrect"]:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        # 答题时间
        cell = ws.cell(row=idx, column=6, value=record["time"])
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border

    # 生成文件名（带时间戳）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"答题记录_{timestamp}.xlsx"
    filepath = os.path.join(export_dir, filename)

    # 保存文件
    wb.save(filepath)

    return filepath

@app.get("/api/export/quiz-records")
async def export_quiz_records(db: Session = Depends(get_db)):
    """导出答题记录为Excel文件"""
    try:
        filepath = export_quiz_records_to_excel(db)
        filename = os.path.basename(filepath)

        return FileResponse(
            path=filepath,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

@app.post("/api/export/auto-save")
async def auto_save_quiz_records(db: Session = Depends(get_db)):
    """自动保存答题记录到Excel（每次答题后调用）"""
    try:
        # 检查数据库中是否有记录
        count = db.query(func.count(QuizRecord.id)).scalar()
        if count == 0:
            return {"success": False, "message": "暂无答题记录"}

        filepath = export_quiz_records_to_excel(db)
        return {
            "success": True,
            "message": "答题记录已保存",
            "filepath": filepath
        }
    except Exception as e:
        return {"success": False, "message": f"保存失败: {str(e)}"}

# ==================== 管理页面 ====================

@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    """管理后台页面"""
    try:
        with open("admin.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        return html_content
    except FileNotFoundError:
        return HTMLResponse(content="<h1>管理页面未找到</h1><p>请确保 admin.html 文件存在于后端目录</p>", status_code=404)

# ==================== 健康检查 ====================

@app.get("/")
async def root():
    """根路径，返回前端页面"""
    try:
        with open("index.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except FileNotFoundError:
        return HTMLResponse(content="<h1>前端页面未找到</h1><p>请确保 index.html 文件存在于后端目录</p>", status_code=404)

@app.get("/health")
async def health_check():
    """健康检查"""
    return {"status": "healthy"}

# ==================== 启动说明 ====================
# 运行命令: uvicorn main:app --reload --port 8000
# API 文档: http://localhost:8000/docs
